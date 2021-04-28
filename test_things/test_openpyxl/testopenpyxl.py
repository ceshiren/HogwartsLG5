#!/usr/bin/env python
# -*- encoding: utf-8 -*-
"""
@File   :   testopenpyxl.py 
@Time   :   2021-04-05 16:13   
@Contact    :      
@Author     :   WG
@Version    :   v 0.1
@Desc  :    openpyxl
"""
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
def write_book():
    wb = Workbook()

    dest_filename = 'empty_book.xlsx'

    ws1 = wb.active

    ws1.title = "range names"
    for row in range(1, 40):
        ws1.append(range(600))
    ws2 = wb.create_sheet(title="Pi")
    ws2['F5'] = 3.14
    ws3 = wb.create_sheet(title="Data")
    for row in range(10, 20):
       for col in range(27, 54):
            _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
    print(ws3['AA10'].value)
    wb.save(filename = dest_filename)
def read_excel():
    wb = load_workbook(filename='empty_book.xlsx')
    sheet_ranges = wb['Data']
    print(sheet_ranges['AA10'].value)
if __name__ == '__main__':
    read_excel()